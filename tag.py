import hashlib
import random
import socket
import json
import threading
from tkinter import *
from tkinter import scrolledtext
from tkinter import messagebox
from tkinter import simpledialog

import openpyxl

class ClientGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Socket Client")
        self.root.geometry("400x300")

        self.log_text = scrolledtext.ScrolledText(self.root, wrap=WORD, width=50, height=15)
        self.log_text.pack(pady=10)

        self.connect_button = Button(self.root, text="Connect to Server", command=self.connect_to_server)
        self.connect_button.pack(pady=10)
        
        self.send_button = Button(self.root, text="", command=self.handle_data, state=DISABLED)
        self.send_button.pack(pady=10)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.client_socket = None
        

    def connect_to_server(self):
        host = '127.0.0.1'
        port = 8000

        self.client_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.client_socket.connect((host, port))

        self.log_message("Connected to the server.")
        self.connect_button.config(state=DISABLED)
        self.send_button.config(state=NORMAL)

        # Start a thread to receive data from the server
        receive_thread = threading.Thread(target=self.handle_data)
        receive_thread.start()
    
    def receive_message(self):
       
        message = self.client_socket.recv(1024).decode('utf-8')
        return message
    
    def send_message_to_server(self,  message):
        
        self.client_socket.send(str.encode(message))

    def handle_successful_authentication(self):
        self.log_message("Authentication successful. You can now start chatting.")

        # Thêm widget Entry và Button cho việc nhập và gửi tin nhắn
        self.message_entry = Entry(self.root)
        self.message_entry.pack(pady=10)
        send_button = Button(self.root, text="Send", command=lambda: self.send_message_to_server( self.message_entry.get()))
        send_button.pack(pady=10)

        # Lặp vô hạn để nhận và gửi tin nhắn
        while True:
           received_message = self.receive_message()
           self.log_message(f"Received message from server: {received_message}")
    def handle_data(self):
        newWin = Tk()
        HashTable = {}
     #But make it invisible

        newWin.withdraw()
        response= self.client_socket.recv(1024)

        name=simpledialog.askstring("Nhập thông tin", response.decode(), parent=newWin)
        self.client_socket.send(str.encode(name))
        # Đợi đến khi hộp thoại được đóng trước khi tiếp tục
        newWin.destroy()
        response=self.client_socket.recv(1024)
        b="Connection Successful"
        a="ENTER UID : "
        file_path = 'datatag.xlsx'
        kts = None
        Trseq = None
        uid=None
        Nt = None
        Nx = None
        AIDt = None
        V1 = None
        def generate_hash(data):
    
            sha256_hash = hashlib.sha256()

   
            sha256_hash.update(str(data).encode('utf-8'))

    
            hash_result = sha256_hash.hexdigest()
            hash_result = int(hash_result, 16)

            return hash_result% 10000000
        
        if response==a.encode('utf-8'):
            newuid=Tk()
            newuid.withdraw()

            uids=simpledialog.askstring("Nhập thông tin", response.decode(), parent=newuid)
            self.client_socket.send(str.encode(uids))
            newuid.destroy()
            response=self.client_socket.recv(1024)
            response=response.decode('utf-8')
            self.log_message(f"{response}")
            
            response=self.client_socket.recv(1024)
            response=response.decode('utf-8')
            ktss=int(response)
            self.log_message(f"Kts: {ktss}")
            response=self.client_socket.recv(1024)
            response=response.decode('utf-8')
            Trseqs=int(response)
            self.log_message(f"Trseq: {response}")
            self.log_message(f"--------------------------------")
            #1
            def write_data_to_excel(file_path,name, uid1, kts,Trseq):
             try:
        
                try:
                   workbook = openpyxl.load_workbook(file_path)
                except FileNotFoundError:
                   workbook = openpyxl.Workbook()

        
                sheet = workbook.active
                x1=sheet["A1"]
                x1.value="Name"
                x2=sheet["B1"]
                x2.value="UID"
                x3=sheet["C1"]
                x3.value="Kts"
                x4=sheet["D1"]
                x4.value="Trseq"
        
                row_number = 1
                while sheet.cell(row=row_number, column=1).value is not None:
                    row_number += 1

        
                sheet.cell(row=row_number, column=1, value=name)
                sheet.cell(row=row_number, column=2, value=uid1)
                sheet.cell(row=row_number, column=3, value=kts)
                sheet.cell(row=row_number, column=4, value=Trseq)

        
                workbook.save(filename=file_path)

        
             except Exception as e:
                print(f"Error: {e}")
            file_path = 'datatag.xlsx'
            write_data_to_excel(file_path,name,uids,ktss,Trseqs)

        else:
         wb=openpyxl.load_workbook('datatag.xlsx')
         sheet=wb.active
         response=response.decode('utf-8')
         self.log_message(f"{response}")
         #1
         name = response[len('Gửi thông tin xác thực'):].split()
         name=name[0]
         row_num = 1  
 
         while row_num <= sheet.max_row:
            cell_value = sheet.cell(row=row_num, column=1).value
    
    
            if cell_value == name:
                kts=sheet.cell(row=row_num,column=3).value
                Trseq=sheet.cell(row=row_num,column=4).value
        
                break
    
            row_num += 1  
 
         self.log_message(f"Trseq: {Trseq}")
         Nt=random.randint(10,100)
         self.log_message(f"Nt: {str(Nt)}")
         Nx=Nt^kts
         self.log_message(f"Nx: {str(Nx)}")

         aidt=(uid,kts,Nt,Trseq)
         AIDt=generate_hash(aidt)
 
         v1=(AIDt,kts,Nx)
         V1=generate_hash(v1)
         print("V1: ",V1)
         print("AIDT: ",AIDt)
         data_to_send = {
            "AIDt": AIDt,
            "Nx": Nx,
            "V1": V1,
            "Trseq":Trseq
         }
         json_data = json.dumps(data_to_send)
         self.client_socket.send(str.encode(json_data))
         data_received = self.client_socket.recv(1024)
         json_data = data_received.decode('utf-8')

         received_data = json.loads(json_data)

         V4=received_data.get('V4')
         Tr=received_data.get('Tr')

         self.log_message(f"V4: {V4}")
         self.log_message(f"Tr: {Tr}")
         v4=(Tr,kts,Nt)
         V4tag=generate_hash(v4)
         if V4tag==V4:
            self.log_message("Xác thực thành công server")
            self.log_message("------------------------------------------------------")

            trseqnew=(Nt,kts)
            Trseqnew=generate_hash(trseqnew)^Tr
            self.log_message(f"Trseqnew: {Trseqnew}")
            
            ktsnew=(kts,Trseqnew)
            Ktsnew=generate_hash(ktsnew)
            self.log_message(f"Ktsnew: {Ktsnew}")
            
   
            row_num = 1  # Bắt đầu từ dòng đầu tiên
            while row_num <= sheet.max_row:
                cell_value = sheet.cell(row=row_num, column=1).value
    
           
                if cell_value == name:
                   sheet.cell(row=row_num,column=3,value=Ktsnew)
                   sheet.cell(row=row_num,column=4,value=Trseqnew)
                   wb.save(filename=file_path)
            
                   break
    
                row_num += 1 
            self.handle_successful_authentication()


    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            if self.client_socket:
                self.client_socket.close()
            self.root.destroy()

    def log_message(self, message):
        self.log_text.insert(END, f"{message}\n")
        self.log_text.yview(END)

if __name__ == "__main__":
    root = Tk()
    app = ClientGUI(root)
    root.mainloop()
