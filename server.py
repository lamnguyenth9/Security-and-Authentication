import socket
import threading
import hashlib
import random
import json
from tkinter import *
from tkinter import scrolledtext
import tkinter
from tkinter import messagebox

import openpyxl
HashTable = {}

class ServerGUI:
    
    def __init__(self, root):
        self.root = root
        self.root.title("Socket Server")
        self.root.geometry("600x400")

        self.log_text = scrolledtext.ScrolledText(self.root, wrap=tkinter.WORD, width=50, height=15)
        self.log_text.pack(pady=10)

        self.start_button = Button(self.root, text="Start Server", command=self.start_server)
        self.start_button.pack(pady=10)

        self.stop_button = Button(self.root, text="Stop Server", command=self.stop_server, state=DISABLED)
        self.stop_button.pack(pady=10)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.server_socket = None
        self.clients = {}

    def start_server(self):
        host = '127.0.0.1'
        port = 8000

        self.server_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.server_socket.bind((host, port))
        self.server_socket.listen(5)

        self.log_message("Server started. Waiting for connections...")

        self.start_button.config(state=DISABLED)
        self.stop_button.config(state=NORMAL)

        # Start a thread to accept client connections
        accept_thread = threading.Thread(target=self.accept_clients)
        accept_thread.start()
  
    def accept_clients(self):
        while True:
            client, address = self.server_socket.accept()
            client_thread = threading.Thread(target=self.handle_client, args=(client,))
            client_thread.start()
    
    def send_message_to_client(self, client_socket, message):
        
        client_socket.send(str.encode(message))
    
    def broadcast_message(self, message):
        
        for client_socket in self.clients.values():
            self.send_message_to_client(client_socket, message)
    
    def receive_message_from_client(self, client_socket):
      
        message = client_socket.recv(1024).decode('utf-8')
        return message
    
    def handle_successful_authentication(self, client_socket):
        self.log_message("Authentication successful. You can now start chatting.")

        # Thêm widget Entry và Button cho việc nhập và gửi tin nhắn
        self.message_entry = Entry(self.root)
        self.message_entry.pack(pady=10)
        send_button = Button(self.root, text="Send", command=lambda: self.send_message_to_client(client_socket, self.message_entry.get()))
        send_button.pack(pady=10)

         # Lặp vô hạn để nhận và gửi tin nhắn
        while True:
           received_message = self.receive_message_from_client(client_socket)
           self.log_message(f"Received message from client: {received_message}")

           # Gửi tin nhắn đến tất cả các client khác
           self.broadcast_message(f"Client {client_socket.getpeername()}: {received_message}")
    def handle_client(self, client_socket):
        def generate_hash(data):
            sha256_hash = hashlib.sha256()
            sha256_hash.update(str(data).encode('utf-8'))
            hash_result = sha256_hash.hexdigest()
            hash_result = int(hash_result, 16)
            return hash_result% 10000000
        
        client_address = client_socket.getpeername()
        self.clients[client_address] = client_socket
        self.log_message(f"New connection")
        
       
        client_socket.send(str.encode('ENTER Tag : '))
        name=client_socket.recv(1024)
        name=name.decode()
        self.log_message(f"Received : {name}")
        def write_data_to_excel(file_path,name, data, data1,data2):
            try:
                try:
                    workbook=openpyxl.load_workbook(file_path)
                except FileNotFoundError:
                    workbook = openpyxl.Workbook()
                sheet = workbook.active
                x1=sheet["A1"]
                x1.value="TagName"
                x2=sheet["B1"]
                x2.value="UID"
        
                x3=sheet["C1"]
                x3.value="Kts"
                x4=sheet["D1"]
                x4.value="Trseq"

                row_number = 1
                while sheet.cell(row=row_number, column=1).value is not None:
                    row_number += 1
                
                sheet.cell(row=row_number, column=1, value=name)
                sheet.cell(row=row_number, column=2, value=data)
                sheet.cell(row=row_number, column=3, value=data1)
                sheet.cell(row=row_number, column=4, value=data2)

                workbook.save(filename=file_path)
            except Exception as e:
              print(f"Error: {e}")
        file_path = 'data.xlsx'

 
        data_to_write = name
        # REGISTERATION PHASE   
        # If new user,  regiter in Hashtable Dictionary 
        if name not in HashTable:
           client_socket.send(str.encode('ENTER UID : ')) # Request uid
           uid=client_socket.recv(1024)

           uid=uid.decode()
           self.log_message(f"Received UID : {uid}")
           HashTable[name]=uid
           client_socket.send(str.encode('Registeration Successful'))
           self.log_message("Registeration Successful")
           
           ns=random.randint(10,100)
           a=(uid,ns)
           kts=generate_hash(a)
           self.log_message(f"Kts: {kts}")
           Trseq=random.randint(10,100)
           self.log_message(f"Trseq: {Trseq}")
           client_socket.send(str(kts).encode('utf-8'))
           client_socket.send(str(Trseq).encode('utf-8'))
           #1
           self.log_message(f"Registered: {name}")
           self.log_message("{:<8}{:<20}".format('USER','UID'))
           for k, v in HashTable.items():
               label, num = k,v
               self.log_message("{:<8} {:<20}".format(label, num))
           write_data_to_excel(file_path, data_to_write,uid,kts,Trseq)
           self.log_message("----------------------------------")
        
        else:
         message=f'Gửi thông tin xác thực {name}'
         self.log_message(f"{message}")
         client_socket.send(str.encode(message))
         #1
         data_received = client_socket.recv(1024)
         json_data = data_received.decode('utf-8')
         received_data = json.loads(json_data)
         AIDt = received_data.get('AIDt')
         Nx = received_data.get('Nx')
         V1 = received_data.get('V1')
         Trseqt=received_data.get('Trseq')
         self.log_message(f"AIDt: {AIDt}")
         self.log_message(f"Nx: {Nx}")
         self.log_message(f"V1: {V1}")
         self.log_message(f"Trseq: {Trseqt}")
         wb=openpyxl.load_workbook('data.xlsx')
         sheet=wb.active
         row_num=1
         while row_num <= sheet.max_row:
           cell_value = sheet.cell(row=row_num, column=1).value
    
           
           if cell_value == name:
             ktss=sheet.cell(row=row_num,column=3).value
             Trseqs=sheet.cell(row=row_num,column=4).value
            
             break
    
           row_num += 1
         if Trseqs==Trseqt:
            self.log_message("Xác thực thẻ thành công")
            self.log_message("--------------------")
            v1s=(AIDt,ktss,Nx)
            V1s=generate_hash(v1s)
            if V1s==V1:
               Nt=ktss^Nx
               print("Nt:" , Nt)
           
               Trseqnew= random.randint(10,100)
               tr=(Nt,ktss)
               Tr=generate_hash(tr)^Trseqnew
               print("Tr: ",Tr)
               v4=(Tr,ktss,Nt)
               V4=generate_hash(v4)
               print("V4",V4)
               ktsnew=(ktss,Trseqnew)
               Ktsnew=generate_hash(ktsnew)
               row_numb=1
               while row_numb <= sheet.max_row:
                  cell_value = sheet.cell(row=row_numb, column=1).value
    
           
                  if cell_value == name:
                    sheet.cell(row=row_numb,column=3,value=Ktsnew)
                    sheet.cell(row=row_numb,column=4,value=Trseqnew)
                    wb.save(filename="data.xlsx")
                    break
                  row_numb+=1
               data_to_send = {
                  "V4": V4,
                  "Tr": Tr     
                }
               json_data = json.dumps(data_to_send)
               client_socket.send(str.encode(json_data))
               self.handle_successful_authentication(client_socket)
            else:
               client_socket.send(str.encode('Authentication failed'))
               client_socket.close()

         else: 
            client_socket.send(str.encode('Authentication failed'))
            client_socket.close()


           




        

    def stop_server(self):
        for client_socket in self.clients.values():
            client_socket.close()

        self.server_socket.close()
        self.log_message("Server stopped.")

        self.start_button.config(state=NORMAL)
        self.stop_button.config(state=DISABLED)

    def on_closing(self):
        if messagebox.askokcancel("Quit", "Do you want to quit?"):
            self.stop_server()
            self.root.destroy()

    def log_message(self, message):
        self.log_text.insert(END, f"{message}\n")
        self.log_text.yview(END)

if __name__ == "__main__":
    root = Tk()
    app = ServerGUI(root)
    root.mainloop()
